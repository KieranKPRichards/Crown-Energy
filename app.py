#!/usr/bin/env python3
"""
Crown Energy Meter Reporting Application
==========================================
DATA SOURCE HIERARCHY:
  Single-meter sites:
    Energy (Peak/Std/OffPeak/Total): BR registers (authoritative)
    Reactive energy: BR register (authoritative)
    Max Demand: BR Max Demand sheet (authoritative)
    Power Factor: calculated from profile data
  Summation sites (multiple incomers):
    Energy total: sum of BR total registers across meters
    Energy TOU split: combined profile TOU ratios scaled to BR total
    Reactive: sum of BR reactive registers across meters
    Max Demand: from combined profile (max half-hourly kVA)
    Power Factor: from combined profile data
"""

import os, sys, json, csv, io, math, shutil, subprocess, calendar, re
from datetime import datetime, timedelta
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
CONFIG_FILE = DATA_DIR / "sites.json"
TARIFF_FILE = DATA_DIR / "tariffs.json"

DEFAULT_TARIFFS = {
    "Megaflex": {
        "type": "tou",
        "service_charge": 198.52, "admin_charge": 19.37,
        "network_demand": 24.17, "transmission_network": 10.25,
        "generation_capacity": 8.09, "network_capacity": 35.98,
        "ancillary_service": 0.0039, "legacy_charge": 0.222,
        "electrification_rural": 0.0494, "affordability_subsidy": 0.0469,
        "energy_rates": {
            "low": {"peak": 2.7678, "standard": 1.5562, "off_peak": 1.1115},
            "high": {"peak": 8.1462, "standard": 2.7910, "off_peak": 1.6879}
        },
        "tou_periods": {
            "low": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                    "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                    "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}},
            "high": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                     "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                     "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}}
        }
    },
    "Miniflex": {
        "type": "tou",
        "service_charge": 198.52, "admin_charge": 19.37,
        "network_demand_rate": 0.0961, "network_capacity": 46.22, "generation_capacity": 8.09,
        "ancillary_service": 0.0039, "legacy_charge": 0.222,
        "electrification_rural": 0.0494, "affordability_subsidy": 0.0469,
        "energy_rates": {
            "low": {"peak": 2.7678, "standard": 1.5562, "off_peak": 1.1115},
            "high": {"peak": 8.1462, "standard": 2.7910, "off_peak": 1.6879}
        },
        "tou_periods": {
            "low": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                    "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                    "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}},
            "high": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                     "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                     "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}}
        }
    },
    "Nightsave Urban": {
        "type": "tou_simple",
        "service_charge": 133.5, "admin_charge": 60.17,
        "network_demand_peak_std": 20.23, "transmission_peak_std": 5.35, "network_access": 10.67,
        "energy_rates": {"low":{"peak":0,"standard":0,"off_peak":0.7319},"high":{"peak":1.7131,"standard":0.9803,"off_peak":0.6589}},
        "tou_periods": {
            "low": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                    "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                    "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}},
            "high": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                     "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                     "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}}
        }
    },
    "Tariff D": {
        "type": "demand", "fixed_charge": 3302.3448, "demand_charge": 140.6756, "network_access": 120.572,
        "energy_rates": {"low":{"peak":3.2813,"standard":2.1007,"off_peak":1.6674},"high":{"peak":3.2813,"standard":2.1007,"off_peak":1.6674}},
        "tou_periods": {
            "low": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                    "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                    "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}},
            "high": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                     "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                     "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}}
        }
    },
    "Tariff E": {
        "type": "demand", "fixed_charge": 5207.0984, "demand_charge": 146.0987, "network_access": 96.6283,
        "energy_rates": {"low":{"peak":3.2813,"standard":2.1007,"off_peak":1.6674},"high":{"peak":3.2813,"standard":2.1007,"off_peak":1.6674}},
        "tou_periods": {
            "low": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                    "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                    "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}},
            "high": {"weekday": {"peak":[(7,10),(18,20)],"standard":[(6,7),(10,18),(20,22)],"off_peak":[(0,6),(22,24)]},
                     "saturday":{"peak":[],"standard":[(7,12),(18,20)],"off_peak":[(0,7),(12,18),(20,24)]},
                     "sunday":  {"peak":[],"standard":[],"off_peak":[(0,24)]}}
        }
    }
}

def load_tariffs():
    if TARIFF_FILE.exists():
        try:
            with open(TARIFF_FILE) as f:
                saved = json.load(f)
            # Merge: use saved values but ensure tou_periods from defaults always present
            merged = {}
            for name, default in DEFAULT_TARIFFS.items():
                if name in saved:
                    t = dict(default)
                    s = saved[name]
                    # Update all non-tou_periods fields from saved
                    for k, v in s.items():
                        if k != 'tou_periods':
                            t[k] = v
                    merged[name] = t
                else:
                    merged[name] = dict(default)
            # Include any custom tariffs not in defaults
            for name, t in saved.items():
                if name not in merged:
                    merged[name] = t
            return merged
        except Exception:
            pass
    return json.loads(json.dumps(DEFAULT_TARIFFS))

def save_tariffs(tariffs):
    with open(TARIFF_FILE, 'w') as f:
        json.dump(tariffs, f, indent=2)

TARIFFS = load_tariffs()

def get_season(dt): return "high" if dt.month in (6,7,8) else "low"
def get_day_type(dt):
    d = dt.weekday()
    return "weekday" if d < 5 else ("saturday" if d == 5 else "sunday")

def classify_tou(dt, tariff_name):
    t = TARIFFS[tariff_name]; s = get_season(dt); d = get_day_type(dt); h = dt.hour
    for p in ["peak","standard","off_peak"]:
        for a,b in t["tou_periods"][s][d].get(p,[]):
            if a <= h < b: return p
    return "off_peak"

def parse_profile_csv(filepath):
    result = {'meter_serial':'','ct_ratio':'','vt_ratio':'','records':[]}
    with open(filepath,'r',encoding='utf-8-sig') as f: lines = f.readlines()
    hdr = False
    for line in lines:
        line = line.strip().replace('\r','')
        if not line: continue
        if line.startswith('Meter Serial:'): result['meter_serial']=line.split(':')[1].strip(); continue
        if line.startswith('Meter - CT:'):
            p = line.split('CT:')[1].strip()
            if 'VT:' in p: c,v=p.split('VT:'); result['ct_ratio']=c.strip(); result['vt_ratio']=v.strip()
            continue
        if line.startswith('Date,Time'): hdr=True; continue
        if line.startswith(',,'): continue
        if hdr:
            parts = line.split(',')
            if len(parts)>=3:
                try:
                    dstr=parts[0].strip(); tstr=parts[1].strip()
                    if tstr=='24:00':
                        dt=datetime.strptime(dstr,"%Y/%m/%d")+timedelta(days=1)
                    else:
                        dt=datetime.strptime(f"{dstr} {tstr}","%Y/%m/%d %H:%M")
                    kwh=float(parts[2].strip()) if parts[2].strip() else 0.0
                    kvarh=float(parts[3].strip()) if len(parts)>3 and parts[3].strip() else 0.0
                    result['records'].append({'datetime':dt,'kwh':kwh,'kvarh':kvarh})
                except: pass
    return result

def parse_billing_xls(filepath):
    import openpyxl
    result = {'meter_serial':'','stack_dates':[],'energy_registers':[],'reactive_register':[],
              'md_mva':[],'md_mva_dates':[],'md_mva_times':[],'md_mw':[],'md_mw_dates':[],'md_mw_times':[]}
    ext = Path(filepath).suffix.lower()
    if ext == '.xls':
        tmp = filepath+'x'
        if not os.path.exists(tmp):
            try: subprocess.run(['libreoffice','--headless','--convert-to','xlsx',filepath,'--outdir',str(Path(filepath).parent)],capture_output=True,timeout=60)
            except: pass
        if os.path.exists(tmp): filepath = tmp
    try: wb = openpyxl.load_workbook(filepath,data_only=True,read_only=True)
    except: return result

    def nums(row):
        v=[]
        for i in range(1,min(len(row),15),2):
            x=row[i]
            if x and str(x).strip() not in ['-','','None','N/A']:
                try: v.append(float(str(x).strip().replace(',','')))
                except: v.append(None)
            else: v.append(None)
        return v
    def strs(row):
        return [str(row[i]).strip() if i<len(row) and row[i] and str(row[i]).strip() not in ['None','N/A'] else '' for i in range(1,min(len(row),15),2)]

    if 'Energy Billing' in wb.sheetnames:
        for row in wb['Energy Billing'].iter_rows(values_only=True):
            if not row[0]: continue
            s=str(row[0]).strip()
            if 'Meter Serial' in s: result['meter_serial']=str(row[1]).strip() if row[1] else ''
            elif 'Running & MER Date' in s: result['stack_dates']=strs(row)
            elif 'Wh' in s and 'Acc-Tot_Imp' in s: result['energy_registers'].append(nums(row))
            elif 'varh' in s.lower() and 'Acc-Exvar' in s: result['reactive_register']=nums(row)

    if 'Max Demand Billing' in wb.sheetnames:
        found_mva=found_mw=False; pending=None
        for row in wb['Max Demand Billing'].iter_rows(values_only=True):
            if not row[0]: continue
            s=str(row[0]).strip()
            if 'VA' in s and 'Non-VABlk' in s and not found_mva:
                result['md_mva']=nums(row); pending='mva_date'; found_mva=True
            elif 'W ' in s and 'Non-Blk' in s and not found_mw:
                result['md_mw']=nums(row); pending='mw_date'; found_mw=True
            elif s=='Date':
                if pending=='mva_date': result['md_mva_dates']=strs(row); pending='mva_time'
                elif pending=='mw_date': result['md_mw_dates']=strs(row); pending='mw_time'
            elif s=='Time':
                if pending=='mva_time': result['md_mva_times']=strs(row); pending=None
                elif pending=='mw_time': result['md_mw_times']=strs(row); pending=None
    wb.close()
    return result

def get_br_month_energy(br, stack_idx=1):
    regs=br['energy_registers']; rr=br['reactive_register']
    if len(regs)<4: return None
    s1,s2=stack_idx,stack_idx+1
    def d(r,i,j):
        if i<len(r) and j<len(r) and r[i] is not None and r[j] is not None:
            return round((r[i]-r[j])*1000,1)
        return None
    pk=d(regs[0],s1,s2); st=d(regs[1],s1,s2); op=d(regs[2],s1,s2); tot=d(regs[3],s1,s2)
    react=d([v if v is not None else 0 for v in rr],s1,s2) if rr else None
    if tot is None: return None
    return {'peak':pk or 0,'standard':st or 0,'off_peak':op or 0,'total':tot,'reactive':react or 0}

def get_br_md(br, stack_idx=1):
    r={}
    ml=br.get('md_mva',[])
    if stack_idx<len(ml) and ml[stack_idx] is not None: r['md_kva']=round(ml[stack_idx]*1000,1)
    wl=br.get('md_mw',[])
    if stack_idx<len(wl) and wl[stack_idx] is not None: r['md_mw']=round(wl[stack_idx],4)
    dl=br.get('md_mva_dates',[]); tl=br.get('md_mva_times',[])
    if stack_idx<len(dl): r['md_date']=dl[stack_idx]
    if stack_idx<len(tl): r['md_time']=tl[stack_idx]
    return r

def analyse_profile(profile, by, bm, tn):
    pk=st=op=tr=0.0; mx=0.0
    for rec in profile['records']:
        dt=rec['datetime']; sdt=dt-timedelta(minutes=30)
        if sdt.year!=by or sdt.month!=bm: continue
        kwh,kvarh=rec['kwh'],rec['kvarh']; tr+=kvarh
        kw,kvar=kwh*2,kvarh*2
        kva=math.sqrt(kw**2+kvar**2) if (kw>0 or kvar>0) else 0
        mx=max(mx,kva)
        p=classify_tou(sdt,tn)
        if p=='peak': pk+=kwh
        elif p=='standard': st+=kwh
        else: op+=kwh
    tot=pk+st+op
    pf=tot/math.sqrt(tot**2+tr**2) if tot>0 else 1.0
    return {'p_pk':round(pk,1),'p_st':round(st,1),'p_op':round(op,1),'p_tot':round(tot,1),'p_react':round(tr,1),
            'pk_r':pk/tot if tot>0 else 0,'st_r':st/tot if tot>0 else 0,'op_r':op/tot if tot>0 else 0,
            'md_kva':round(mx,2),'pf':round(pf,6)}

def sum_profiles(profs):
    if len(profs)==1: return profs[0]
    c={}
    for p in profs:
        for r in p['records']:
            k=r['datetime']
            if k not in c: c[k]={'datetime':k,'kwh':0.0,'kvarh':0.0}
            c[k]['kwh']+=r['kwh']; c[k]['kvarh']+=r['kvarh']
    return {'meter_serial':'+'.join(p['meter_serial'] for p in profs),'ct_ratio':'','vt_ratio':'',
            'records':sorted(c.values(),key=lambda r:r['datetime'])}

def assemble_energy_data(brs, profs, is_sum, by, bm, tn):
    cp = sum_profiles(profs) if len(profs)>1 else profs[0]
    pa = analyse_profile(cp, by, bm, tn)

    if not is_sum and len(brs)==1:
        be=get_br_month_energy(brs[0]); bmd=get_br_md(brs[0])
        if be:
            return {'peak':be['peak'],'standard':be['standard'],'off_peak':be['off_peak'],
                    'total':be['total'],'peak_std':be['peak']+be['standard'],
                    'reactive':be['reactive'],'max_demand_kva':bmd.get('md_kva',pa['md_kva']),
                    'md_date':bmd.get('md_date',''),'md_time':bmd.get('md_time',''),
                    'power_factor':pa['pf'],'data_source':'billing_registers'}

    if is_sum:
        bt=br_r=0; ok=True
        for b in brs:
            e=get_br_month_energy(b)
            if e: bt+=e['total']; br_r+=e['reactive']
            else: ok=False
        if ok and bt>0:
            return {'peak':round(bt*pa['pk_r'],1),'standard':round(bt*pa['st_r'],1),
                    'off_peak':round(bt*pa['op_r'],1),'total':round(bt,1),
                    'peak_std':round(bt*(pa['pk_r']+pa['st_r']),1),
                    'reactive':round(br_r,1),'max_demand_kva':pa['md_kva'],
                    'md_date':'','md_time':'','power_factor':pa['pf'],
                    'data_source':'billing_total_plus_profile_ratios'}

    return {'peak':pa['p_pk'],'standard':pa['p_st'],'off_peak':pa['p_op'],'total':pa['p_tot'],
            'peak_std':pa['p_pk']+pa['p_st'],'reactive':pa['p_react'],
            'max_demand_kva':pa['md_kva'],'md_date':'','md_time':'',
            'power_factor':pa['pf'],'data_source':'profile_only'}

def calculate_bill(energy, tn, nmd, ucap, days, bm, r12m=None):
    t=TARIFFS[tn]; s=get_season(datetime(2026,bm,1)); rates=t['energy_rates'][s]; b={}
    if tn=='Megaflex':
        b['service_charge']=round(t['service_charge']*days,2); b['admin_charge']=round(t['admin_charge']*days,2)
        b['network_demand']=round(t['network_demand']*energy['max_demand_kva'],2)
        b['transmission_network']=round(t['transmission_network']*ucap,2)
        b['generation_capacity']=round(t['generation_capacity']*ucap,2)
        b['network_capacity']=round(t['network_capacity']*ucap,2)
        b['energy_peak']=round(rates['peak']*energy['peak'],2)
        b['energy_standard']=round(rates['standard']*energy['standard'],2)
        b['energy_off_peak']=round(rates['off_peak']*energy['off_peak'],2)
        tk=energy['total']
        b['ancillary_service']=round(t['ancillary_service']*tk,2)
        b['legacy_charge']=round(t['legacy_charge']*tk,2)
        b['electrification_rural']=round(t['electrification_rural']*tk,2)
        b['affordability_subsidy']=round(t['affordability_subsidy']*tk,2)
        b['reactive_energy']=0
    elif tn=='Miniflex':
        b['service_charge']=round(t['service_charge']*days,2); b['admin_charge']=round(t['admin_charge']*days,2)
        b['network_demand']=round(t['network_demand_rate']*energy['peak_std'],2)
        b['network_capacity']=round(t['network_capacity']*nmd,2)
        b['generation_capacity']=round(t['generation_capacity']*nmd,2)
        b['energy_peak']=round(rates['peak']*energy['peak'],2)
        b['energy_standard']=round(rates['standard']*energy['standard'],2)
        b['energy_off_peak']=round(rates['off_peak']*energy['off_peak'],2)
        tk=energy['total']
        b['ancillary_service']=round(t['ancillary_service']*tk,2)
        b['legacy_charge']=round(t['legacy_charge']*tk,2)
        b['electrification_rural']=round(t['electrification_rural']*tk,2)
        b['affordability_subsidy']=round(t['affordability_subsidy']*tk,2)
        b['reactive_energy']=0
    elif t['type']=='demand':
        b['fixed_charge']=round(t['fixed_charge'],2)
        b['demand_charge']=round(t['demand_charge']*energy['max_demand_kva'],2)
        rm=r12m if r12m else energy['max_demand_kva']
        b['network_access']=round(t['network_access']*rm,2)
        b['energy_peak']=round(rates['peak']*energy['peak'],2)
        b['energy_standard']=round(rates['standard']*energy['standard'],2)
        b['energy_off_peak']=round(rates['off_peak']*energy['off_peak'],2)
    b['total_energy']=round(b.get('energy_peak',0)+b.get('energy_standard',0)+b.get('energy_off_peak',0),2)
    sub=sum(v for k,v in b.items() if isinstance(v,(int,float)) and k!='total_energy')
    b['subtotal_excl_vat']=round(sub,2); b['vat']=round(sub*0.15,2); b['total_incl_vat']=round(sub*1.15,2)
    return b

def parse_filename_metadata(filename):
    """Extract meter number and billing period from an Enermax filename.
    Handles patterns like:
      BR14140031_2026-01-01_2026-01-31.xlsx
      PR14140031_20260101_20260131.csv
      BR 14140031 2026-01-01 2026-01-31.xls
    Returns {'meter_number': str, 'period': 'YYYY-MM'} or None.
    """
    stem = Path(filename).stem
    m = re.search(r'(?:BR|PR)[_\s]?(\d{5,12})[_\s](\d{4}[-/]\d{2}[-/]\d{2}|\d{8})', stem, re.IGNORECASE)
    if not m:
        return None
    meter_num = m.group(1)
    date_str = m.group(2)
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d']:
        try:
            dt = datetime.strptime(date_str, fmt)
            return {'meter_number': meter_num, 'period': dt.strftime('%Y-%m')}
        except ValueError:
            pass
    return None

def extract_meter_serial(filepath):
    """Extract meter serial number from a BR (.xls/.xlsx) or PR (.csv) file."""
    fp = str(filepath)
    ext = Path(fp).suffix.lower()
    if ext == '.csv':
        try:
            with open(fp, 'r', encoding='utf-8-sig') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('Meter Serial:'):
                        return line.split(':', 1)[1].strip()
        except Exception:
            pass
    elif ext in ('.xls', '.xlsx'):
        parsed = parse_billing_xls(fp)
        return parsed.get('meter_serial', '')
    return ''

from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import UniqueConstraint

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'crown-energy-dev-key')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', f'sqlite:///{DATA_DIR}/energy.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'connect_args': {'timeout': 30}}
db = SQLAlchemy(app)


# ---------------------------------------------------------------------------
# SQLAlchemy Models
# ---------------------------------------------------------------------------

class Site(db.Model):
    __tablename__ = 'sites'
    id = db.Column(db.String(20), primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    plant = db.Column(db.String(200), default='')
    tariff = db.Column(db.String(100), nullable=False)
    supply_authority = db.Column(db.String(100), default='Eskom')
    account_number = db.Column(db.String(100), default='')
    nmd_kva = db.Column(db.Float, default=0.0)
    utilised_capacity_kva = db.Column(db.Float, default=0.0)
    voltage = db.Column(db.String(50), default='')
    is_summation = db.Column(db.Boolean, default=False)
    meters = db.relationship('Meter', back_populates='site', cascade='all, delete-orphan', order_by='Meter.id')
    reports = db.relationship('Report', back_populates='site', cascade='all, delete-orphan', order_by='Report.id')


class Meter(db.Model):
    __tablename__ = 'meters'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    site_id = db.Column(db.String(20), db.ForeignKey('sites.id'), nullable=False)
    label = db.Column(db.String(200), default='')
    meter_number = db.Column(db.String(200), default='')
    billing_file = db.Column(db.String(500), default='')
    profile_file = db.Column(db.String(500), default='')
    site = db.relationship('Site', back_populates='meters')
    file_periods = db.relationship('FilePeriod', back_populates='meter', cascade='all, delete-orphan', order_by='FilePeriod.period')

    @property
    def file_periods_dict(self):
        return {fp.period: {'billing_file': fp.billing_file or '', 'profile_file': fp.profile_file or ''}
                for fp in self.file_periods}


class FilePeriod(db.Model):
    __tablename__ = 'file_periods'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    meter_id = db.Column(db.Integer, db.ForeignKey('meters.id'), nullable=False)
    period = db.Column(db.String(7), nullable=False)
    billing_file = db.Column(db.String(500), default='')
    profile_file = db.Column(db.String(500), default='')
    meter = db.relationship('Meter', back_populates='file_periods')
    __table_args__ = (UniqueConstraint('meter_id', 'period', name='uq_meter_period'),)


class Report(db.Model):
    __tablename__ = 'reports'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    site_id = db.Column(db.String(20), db.ForeignKey('sites.id'), nullable=False)
    generated = db.Column(db.String(30))
    billing_period = db.Column(db.String(7))
    billing_month_name = db.Column(db.String(50))
    days = db.Column(db.Integer)
    tariff = db.Column(db.String(100))
    season = db.Column(db.String(10))
    is_summation = db.Column(db.Boolean)
    num_meters = db.Column(db.Integer)
    nmd_kva = db.Column(db.Float)
    utilised_capacity_kva = db.Column(db.Float)
    data_source = db.Column(db.String(100))
    energy_json = db.Column(db.Text)
    bill_json = db.Column(db.Text)
    files_used_json = db.Column(db.Text)
    site = db.relationship('Site', back_populates='reports')

    @property
    def energy(self):
        return json.loads(self.energy_json or '{}')

    @property
    def bill(self):
        return json.loads(self.bill_json or '{}')

    @property
    def files_used(self):
        return json.loads(self.files_used_json or '[]')


# ---------------------------------------------------------------------------
# JSON migration helper
# ---------------------------------------------------------------------------

def migrate_from_json():
    """Migrate data/sites.json to the DB if DB is empty and JSON exists."""
    if not CONFIG_FILE.exists():
        return
    if Site.query.count() > 0:
        return
    try:
        with open(CONFIG_FILE) as f:
            sites_data = json.load(f)
    except Exception:
        return

    for sd in sites_data:
        site = Site(
            id=sd['id'],
            name=sd['name'],
            plant=sd.get('plant', ''),
            tariff=sd['tariff'],
            supply_authority=sd.get('supply_authority', 'Eskom'),
            account_number=sd.get('account_number', ''),
            nmd_kva=float(sd.get('nmd_kva', 0)),
            utilised_capacity_kva=float(sd.get('utilised_capacity_kva', 0)),
            voltage=sd.get('voltage', ''),
            is_summation=bool(sd.get('is_summation', False)),
        )
        db.session.add(site)
        db.session.flush()

        for md in sd.get('meters', []):
            meter = Meter(
                site_id=site.id,
                label=md.get('label', ''),
                meter_number=md.get('meter_number', md.get('label', '')),
                billing_file=md.get('billing_file', ''),
                profile_file=md.get('profile_file', ''),
            )
            db.session.add(meter)
            db.session.flush()

            for period, pf in md.get('file_periods', {}).items():
                fp = FilePeriod(
                    meter_id=meter.id,
                    period=period,
                    billing_file=pf.get('billing_file', ''),
                    profile_file=pf.get('profile_file', ''),
                )
                db.session.add(fp)

        for rd in sd.get('reports', []):
            report = Report(
                site_id=site.id,
                generated=rd.get('generated', ''),
                billing_period=rd.get('billing_period', ''),
                billing_month_name=rd.get('billing_month_name', ''),
                days=rd.get('days'),
                tariff=rd.get('tariff', ''),
                season=rd.get('season', ''),
                is_summation=bool(rd.get('is_summation', False)),
                num_meters=rd.get('num_meters'),
                nmd_kva=rd.get('nmd_kva'),
                utilised_capacity_kva=rd.get('utilised_capacity_kva'),
                data_source=rd.get('data_source', ''),
                energy_json=json.dumps(rd.get('energy', {})),
                bill_json=json.dumps(rd.get('bill', {})),
                files_used_json=json.dumps(rd.get('files_used', [])),
            )
            db.session.add(report)

    db.session.commit()


# Initialise DB tables and run migration
with app.app_context():
    db.create_all()
    migrate_from_json()


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------

@app.route('/health')
def health():
    return jsonify({'status': 'ok'}), 200


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def prev_month_period():
    today = datetime.today()
    first = today.replace(day=1)
    last_month = first - timedelta(days=1)
    return last_month.strftime('%Y-%m')


def get_period_status(site, period_key):
    if not site.meters:
        return 'empty'
    has_b = has_p = 0
    n = len(site.meters)
    for m in site.meters:
        fp = next((f for f in m.file_periods if f.period == period_key), None)
        if (fp and fp.billing_file) or m.billing_file:
            has_b += 1
        if (fp and fp.profile_file) or m.profile_file:
            has_p += 1
    if Report.query.filter_by(site_id=site.id, billing_period=period_key).first():
        return 'reported'
    if has_b == n and has_p == n:
        return 'complete'
    if has_b > 0 or has_p > 0:
        return 'partial'
    return 'empty'


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    sites = Site.query.all()
    pp = prev_month_period()
    statuses = {s.id: get_period_status(s, pp) for s in sites}
    return render_template('index.html', sites=sites, tariffs=list(TARIFFS.keys()),
                           prev_period=pp, site_statuses=statuses)


@app.route('/add_site', methods=['POST'])
def add_site():
    meter_numbers = [n.strip() for n in request.form.get('meter_numbers', '').splitlines() if n.strip()]
    site = Site(
        id=datetime.now().strftime('%Y%m%d%H%M%S'),
        name=request.form['name'],
        plant=request.form.get('plant', ''),
        tariff=request.form['tariff'],
        supply_authority=request.form.get('supply_authority', 'Eskom'),
        account_number=request.form.get('account_number', ''),
        nmd_kva=float(request.form.get('nmd_kva', 0)),
        utilised_capacity_kva=float(request.form.get('utilised_capacity_kva', 0)),
        voltage=request.form.get('voltage', ''),
        is_summation='is_summation' in request.form,
    )
    db.session.add(site)
    db.session.flush()
    for num in meter_numbers:
        meter = Meter(site_id=site.id, label=num, meter_number=num)
        db.session.add(meter)
    db.session.commit()
    flash(f'Site "{site.name}" added.', 'success')
    return redirect(url_for('index'))


@app.route('/delete_site/<sid>')
def delete_site(sid):
    site = db.get_or_404(Site, sid)
    db.session.delete(site)
    db.session.commit()
    d = DATA_DIR / sid
    if d.exists():
        shutil.rmtree(d)
    flash('Site deleted.', 'info')
    return redirect(url_for('index'))


@app.route('/site/<sid>')
def site_detail(sid):
    site = db.get_or_404(Site, sid)
    return render_template('site.html', site=site, tariffs=list(TARIFFS.keys()))


@app.route('/upload/<sid>', methods=['POST'])
def upload_files(sid):
    site = db.get_or_404(Site, sid)
    d = DATA_DIR / sid
    d.mkdir(exist_ok=True)
    label = request.form.get('meter_label', 'Meter')
    meter = Meter(site_id=site.id, label=label, meter_number=label)
    db.session.add(meter)
    db.session.flush()

    for ft in ['billing_file', 'profile_file']:
        f = request.files.get(ft)
        if f and f.filename:
            f.save(str(d / f.filename))
            if ft == 'billing_file':
                meter.billing_file = f.filename
            else:
                meter.profile_file = f.filename
            meta = parse_filename_metadata(f.filename)
            if meta and meta.get('period'):
                p = meta['period']
                fp = FilePeriod.query.filter_by(meter_id=meter.id, period=p).first()
                if not fp:
                    fp = FilePeriod(meter_id=meter.id, period=p)
                    db.session.add(fp)
                    db.session.flush()
                if ft == 'billing_file':
                    fp.billing_file = f.filename
                else:
                    fp.profile_file = f.filename

    db.session.commit()
    flash(f'Files uploaded for {meter.label}.', 'success')
    return redirect(url_for('site_detail', sid=sid))


@app.route('/remove_meter/<sid>/<int:meter_id>')
def remove_meter(sid, meter_id):
    meter = db.get_or_404(Meter, meter_id)
    if meter.site_id != sid:
        flash('Not found.', 'error')
        return redirect(url_for('site_detail', sid=sid))
    db.session.delete(meter)
    db.session.commit()
    return redirect(url_for('site_detail', sid=sid))


@app.route('/generate_report/<sid>', methods=['POST'])
def generate_report(sid):
    site = db.get_or_404(Site, sid)
    by = int(request.form['billing_year'])
    bm = int(request.form['billing_month'])
    period_key = f"{by}-{bm:02d}"
    d = DATA_DIR / sid
    profs = []; brs = []; files_used = []

    for m in site.meters:
        mn = m.meter_number or m.label or ''
        fp = next((f for f in m.file_periods if f.period == period_key), None)
        bf = (fp.billing_file if fp and fp.billing_file else None) or m.billing_file or ''
        pf = (fp.profile_file if fp and fp.profile_file else None) or m.profile_file or ''
        match = 'period' if fp else 'fallback'
        files_used.append({'meter': mn, 'billing_file': bf, 'profile_file': pf, 'match': match})
        if pf:
            pp = d / pf
            if pp.exists():
                profs.append(parse_profile_csv(str(pp)))
        if bf:
            bp = d / bf
            if bp.exists():
                brs.append(parse_billing_xls(str(bp)))

    if not profs and not brs:
        flash('No data files for the selected period. Upload BR and/or PR files.', 'error')
        return redirect(url_for('site_detail', sid=sid))
    if not profs:
        flash('Profile (PR) data required for MD/PF.', 'error')
        return redirect(url_for('site_detail', sid=sid))

    energy = assemble_energy_data(brs, profs, site.is_summation, by, bm, site.tariff)
    days = calendar.monthrange(by, bm)[1]
    bill = calculate_bill(energy, site.tariff, site.nmd_kva or 0, site.utilised_capacity_kva or 0, days, bm)

    report = Report(
        site_id=site.id,
        generated=datetime.now().isoformat(),
        billing_period=period_key,
        billing_month_name=datetime(by, bm, 1).strftime('%B %Y'),
        days=days,
        tariff=site.tariff,
        season=get_season(datetime(by, bm, 1)),
        is_summation=site.is_summation,
        num_meters=len(profs),
        nmd_kva=site.nmd_kva,
        utilised_capacity_kva=site.utilised_capacity_kva,
        data_source=energy.get('data_source', ''),
        energy_json=json.dumps(energy),
        bill_json=json.dumps(bill),
        files_used_json=json.dumps(files_used),
    )
    db.session.add(report)
    db.session.commit()
    flash(f'Report generated for {report.billing_month_name}.', 'success')
    return redirect(url_for('view_report', sid=sid, ri=report.id))


@app.route('/report/<sid>/<int:ri>')
def view_report(sid, ri):
    report = db.get_or_404(Report, ri)
    if report.site_id != sid:
        flash('Not found.', 'error')
        return redirect(url_for('index'))
    site = db.get_or_404(Site, sid)
    return render_template('report.html', site=site, report=report, report_idx=report.id)


@app.route('/export_report/<sid>/<int:ri>')
def export_report(sid, ri):
    report = db.get_or_404(Report, ri)
    if report.site_id != sid:
        return "Not found", 404
    site = db.get_or_404(Site, sid)
    o = io.StringIO(); w = csv.writer(o)
    w.writerow(['Crown Energy Report']); w.writerow(['Site', site.name])
    w.writerow(['Period', report.billing_month_name]); w.writerow(['Tariff', report.tariff])
    w.writerow(['Source', report.data_source or '']); w.writerow([])
    w.writerow(['ENERGY'])
    energy = report.energy
    for k in ['peak', 'standard', 'off_peak', 'total', 'reactive', 'max_demand_kva', 'power_factor']:
        w.writerow([k, energy.get(k, '')])
    w.writerow([]); w.writerow(['BILL'])
    for k, v in report.bill.items():
        w.writerow([k.replace('_', ' ').title(), f'R {v:,.2f}'])
    o.seek(0)
    return send_file(io.BytesIO(o.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True,
                     download_name=f"report_{site.name}_{report.billing_period}.csv")


@app.route('/update_site/<sid>', methods=['POST'])
def update_site(sid):
    site = db.get_or_404(Site, sid)
    for k in ['name', 'plant', 'tariff', 'supply_authority', 'account_number', 'voltage']:
        setattr(site, k, request.form.get(k, getattr(site, k, '') or ''))
    site.nmd_kva = float(request.form.get('nmd_kva', 0))
    site.utilised_capacity_kva = float(request.form.get('utilised_capacity_kva', 0))
    site.is_summation = 'is_summation' in request.form

    new_numbers = [n.strip() for n in request.form.get('meter_numbers', '').splitlines() if n.strip()]
    if new_numbers:
        old_by_num = {m.meter_number or m.label: m for m in site.meters}
        new_meters = []
        for num in new_numbers:
            if num in old_by_num:
                m = old_by_num[num]
                m.meter_number = num
                new_meters.append(m)
            else:
                new_m = Meter(site_id=site.id, label=num, meter_number=num)
                db.session.add(new_m)
                new_meters.append(new_m)
        # Delete meters not in the new list
        new_ids = {id(m) for m in new_meters}
        for m in site.meters:
            if id(m) not in new_ids and m not in new_meters:
                db.session.delete(m)

    db.session.commit()
    flash('Updated.', 'success')
    return redirect(url_for('site_detail', sid=sid))


@app.route('/delete_report/<sid>/<int:rid>')
def delete_report(sid, rid):
    report = db.get_or_404(Report, rid)
    if report.site_id != sid:
        flash('Not found.', 'error')
        return redirect(url_for('site_detail', sid=sid))
    db.session.delete(report)
    db.session.commit()
    flash('Report deleted.', 'info')
    return redirect(url_for('site_detail', sid=sid))


@app.route('/clear_period_file/<int:meter_id>/<period>/<filetype>')
def clear_period_file(meter_id, period, filetype):
    fp = FilePeriod.query.filter_by(meter_id=meter_id, period=period).first_or_404()
    meter = db.get_or_404(Meter, meter_id)
    if filetype == 'billing_file':
        fp.billing_file = ''
    elif filetype == 'profile_file':
        fp.profile_file = ''
    if not fp.billing_file and not fp.profile_file:
        db.session.delete(fp)
    db.session.commit()
    return redirect(url_for('site_detail', sid=meter.site_id))


@app.route('/report_preview/<sid>')
def report_preview(sid):
    site = Site.query.get(sid)
    if not site:
        return jsonify({'error': 'not found'}), 404
    try:
        by = int(request.args.get('year', datetime.now().year))
        bm = int(request.args.get('month', datetime.now().month))
    except ValueError:
        return jsonify({'error': 'invalid params'}), 400
    period_key = f"{by}-{bm:02d}"
    d = DATA_DIR / sid
    meters_out = []
    for m in site.meters:
        mn = m.meter_number or m.label or ''
        fp = next((f for f in m.file_periods if f.period == period_key), None)
        bf = (fp.billing_file if fp and fp.billing_file else None) or m.billing_file or ''
        pf = (fp.profile_file if fp and fp.profile_file else None) or m.profile_file or ''
        match = 'period' if fp else 'fallback'
        meters_out.append({
            'meter': mn,
            'billing_file': bf,
            'billing_ok': bool(bf) and (d / bf).exists() if bf else False,
            'profile_file': pf,
            'profile_ok': bool(pf) and (d / pf).exists() if pf else False,
            'match': match,
        })
    return jsonify({'period': period_key, 'meters': meters_out})


@app.route('/bulk_upload', methods=['GET', 'POST'])
def bulk_upload():
    if request.method == 'GET':
        return render_template('bulk_upload.html', results=None)
    files = request.files.getlist('files')
    if not files or all(not f.filename for f in files):
        flash('No files selected.', 'error')
        return render_template('bulk_upload.html', results=None)

    sites = Site.query.all()
    # Build lookup: meter_number -> (site_obj, meter_obj)
    lookup = {}
    for site in sites:
        for meter in site.meters:
            mn = (meter.meter_number or meter.label or '').strip()
            if mn:
                lookup[mn] = (site, meter)

    tmp_dir = DATA_DIR / '_tmp_upload'
    tmp_dir.mkdir(exist_ok=True)
    results = []

    for f in files:
        if not f.filename:
            continue
        fname = f.filename
        ext = Path(fname).suffix.lower()
        tmp_path = tmp_dir / fname
        f.save(str(tmp_path))

        period = None
        meta = parse_filename_metadata(fname)
        if meta and meta['meter_number'] in lookup:
            serial = meta['meter_number']
            period = meta['period']
        else:
            serial = extract_meter_serial(str(tmp_path))
            if meta:
                period = meta.get('period')

        if not serial or serial not in lookup:
            results.append({'file': fname, 'serial': serial or '—', 'period': period or '—',
                            'status': 'unmatched', 'site': '—', 'type': '—'})
            continue

        site, meter = lookup[serial]
        dest_dir = DATA_DIR / site.id
        dest_dir.mkdir(exist_ok=True)
        dest = dest_dir / fname
        shutil.move(str(tmp_path), str(dest))

        ftype = 'Profile (PR)' if ext == '.csv' else 'Billing (BR)'
        file_field = 'profile_file' if ext == '.csv' else 'billing_file'

        # Update legacy fallback field on meter
        setattr(meter, file_field, fname)

        if period:
            fp = FilePeriod.query.filter_by(meter_id=meter.id, period=period).first()
            if not fp:
                fp = FilePeriod(meter_id=meter.id, period=period)
                db.session.add(fp)
                db.session.flush()
            setattr(fp, file_field, fname)

        results.append({'file': fname, 'serial': serial, 'period': period or '—',
                        'status': 'assigned', 'site': site.name, 'type': ftype})

    if tmp_dir.exists():
        for p in tmp_dir.iterdir():
            p.unlink(missing_ok=True)
        tmp_dir.rmdir()

    db.session.commit()
    assigned = sum(1 for r in results if r['status'] == 'assigned')
    flash(f'{assigned} of {len(results)} files assigned to sites.', 'success' if assigned else 'info')
    return render_template('bulk_upload.html', results=results)


@app.route('/overview')
def overview():
    sites = Site.query.all()
    today = datetime.today()
    months = []
    y, m = today.year, today.month
    for _ in range(12):
        months.insert(0, f"{y}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    grid = {s.id: {p: get_period_status(s, p) for p in months} for s in sites}
    return render_template('overview.html', sites=sites, months=months, grid=grid)


@app.route('/tariff_editor')
def tariff_editor():
    global TARIFFS
    TARIFFS = load_tariffs()
    return render_template('tariff_editor.html', tariffs=TARIFFS, defaults=DEFAULT_TARIFFS)


@app.route('/update_tariff/<tname>', methods=['POST'])
def update_tariff(tname):
    global TARIFFS
    TARIFFS = load_tariffs()
    if tname not in TARIFFS:
        flash('Tariff not found.', 'error')
        return redirect(url_for('tariff_editor'))
    t = TARIFFS[tname]
    for key in request.form:
        val = request.form[key].strip()
        if not val:
            continue
        try:
            fval = float(val)
        except ValueError:
            continue
        if key.startswith('energy_'):
            parts = key.split('_', 2)  # energy, low/high, peak/standard/off_peak
            if len(parts) == 3:
                season, period = parts[1], parts[2]
                if season in t.get('energy_rates', {}) and period in t['energy_rates'].get(season, {}):
                    t['energy_rates'][season][period] = fval
        else:
            if key in t:
                t[key] = fval
    save_tariffs(TARIFFS)
    flash(f'Tariff "{tname}" updated.', 'success')
    return redirect(url_for('tariff_editor'))


@app.route('/reset_tariff/<tname>')
def reset_tariff(tname):
    global TARIFFS
    TARIFFS = load_tariffs()
    if tname in DEFAULT_TARIFFS:
        TARIFFS[tname] = json.loads(json.dumps(DEFAULT_TARIFFS[tname]))
        save_tariffs(TARIFFS)
        flash(f'Tariff "{tname}" reset to defaults.', 'success')
    return redirect(url_for('tariff_editor'))


@app.route('/reset_all_tariffs')
def reset_all_tariffs():
    global TARIFFS
    TARIFFS = json.loads(json.dumps(DEFAULT_TARIFFS))
    save_tariffs(TARIFFS)
    flash('All tariffs reset to defaults.', 'success')
    return redirect(url_for('tariff_editor'))


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"\n{'='*60}\n  Crown Energy Meter Reporting\n  http://localhost:{port}\n{'='*60}\n")
    app.run(host='0.0.0.0', port=port, debug=True)
